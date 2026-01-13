import os
import pandas as pd
from Bio import SeqIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Define the directory paths
input_dir = r"C:\Users\iulia\PycharmProjects\pythonProject"
output_dir = r"C:\Users\iulia\PycharmProjects\pythonProject"

# Leucine and Serine codons
leucine_codons = ["TTA", "TTG", "CTT", "CTC", "CTA", "CTG"]
serine_codons = ["TCT", "TCC", "TCA", "TCG", "AGT", "AGC"]

# Leucine classifications
leucine_classifications = {
    "1-2-stop": ["TTA", "TTG"],
    "2-2-stop": ["CTA", "CTG"],
    "No-stop": ["CTT", "CTC"]
}

# Serine classifications
serine_classifications = {
    "1-2-stop": ["TCA", "TCG"],
    "2-2-stop": ["TCT", "TCC"],
    "No-stop": ["AGT", "AGC"]
}


def translate_sequence(seq):
    """Translate nucleotide sequence to amino acids."""
    try:
        return str(seq.translate())
    except Exception as e:
        print(f"Error in translating sequence: {e}")
        return ""


def classify_and_count_codons(nucleotide_seq, codons, classifications):
    """Classify and count codons in the sequence."""
    codon_counts = {key: 0 for key in classifications.keys()}
    total_count = 0

    for i in range(0, len(nucleotide_seq) - 2, 3):
        codon = nucleotide_seq[i:i + 3]
        if (codon in codons) and (len(codon) == 3):
            total_count += 1
            for classification, codon_list in classifications.items():
                if codon in codon_list:
                    codon_counts[classification] += 1

    return total_count, codon_counts


def process_genbank_file(genbank_file):
    """Process the GenBank file and save results."""
    try:
        records = SeqIO.parse(genbank_file, "genbank")
    except Exception as e:
        print(f"Error reading GenBank file: {e}")
        return

    results = []

    for record in records:
        nucleotide_seq = str(record.seq)
        amino_acid_seq = translate_sequence(record.seq)

        total_leucines, leucine_counts = classify_and_count_codons(nucleotide_seq, leucine_codons, leucine_classifications)
        total_serines, serine_counts = classify_and_count_codons(nucleotide_seq, serine_codons, serine_classifications)
        total_amino_acids = len(amino_acid_seq)

        result = {
            "Record ID": record.id,
            "Nucleotide Sequence": nucleotide_seq,
            "Amino Acid Sequence": amino_acid_seq,
            "Total Amino Acids": total_amino_acids,
            "Total Leucines": total_leucines,
            "Leucine 1-2-stop": leucine_counts["1-2-stop"],
            "Leucine 2-2-stop": leucine_counts["2-2-stop"],
            "Leucine No-stop": leucine_counts["No-stop"],
            "Total Serines": total_serines,
            "Serine 1-2-stop": serine_counts["1-2-stop"],
            "Serine 2-2-stop": serine_counts["2-2-stop"],
            "Serine No-stop": serine_counts["No-stop"],
            "Issue": len(nucleotide_seq) % 3 != 0  # Add an issue flag
        }
        results.append(result)

    # Convert results to DataFrame
    df = pd.DataFrame(results)

    # Export to Excel
    output_file = os.path.join(output_dir, "Leucine_and_Serine_analysis.xlsx")
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Analysis Results', index=False)
        print(f"Analysis complete. Results saved to {output_file}")

        # Load the workbook and the specific sheet
        workbook = load_workbook(output_file)
        sheet = workbook['Analysis Results']

        # Define the fill for highlighting
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Apply highlighting based on the 'Issue' column
        for row in range(2, len(df) + 2):  # DataFrame rows start from 0, Excel rows from 1, header row is 1
            issue = sheet[f"O{row}"].value  # Column 'O' for 'Issue'
            if issue:
                for col in range(1, len(df.columns) + 1):
                    sheet.cell(row=row, column=col).fill = fill

        # Save the updated workbook
        workbook.save(output_file)

    except Exception as e:
        print(f"Error saving Excel file: {e}")


def find_genbank_file():
    """Find the first .gb or .gbk file in the input directory."""
    for filename in os.listdir(input_dir):
        if filename.endswith(".gb") or filename.endswith(".gbk"):
            return os.path.join(input_dir, filename)
    return None


if __name__ == "__main__":
    genbank_file = find_genbank_file()
    if genbank_file:
        print(f"Processing GenBank file: {genbank_file}")
        process_genbank_file(genbank_file)
    else:
        print("No GenBank file found in the directory.")
